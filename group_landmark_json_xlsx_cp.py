import itertools
import json
from openpyxl import load_workbook

""" 
group landmark 를 정의 하는 과정 에서 landmark 가 올바른지 확인 하는 코드
json 에서만 존재 하는 landmark, xlsx 에서만 존재 하는 landmark 가 출력됨
"""

# json 과 xlsx 파일 입력
json_path = r'./group_points_preShin.json'
xlsx_path = r'./sample.xlsx'

# xlsx_path = r'\\3dons_server\Share\ON3D, ON3D_S Analysis 최신 엑셀\Dr_Cho_Analysis_230321(prp 삭제).xlsx'
# json_path = r'./group_points_preShin.json'


def load_json(path: str):
    """
    json group landmark 중복 제거 후 list 에 저장
    train 이나 exporter 에서 landmark 를 넣을 때 편하도록 landmark list 와 한줄씩 landmark num 를 출력
    """

    with open(path, 'r') as f:  # json 파일 불러와 읽기 : dict
        json_data = json.load(f)

    landmark_list = []

    list2 = list(itertools.chain(*json_data.values()))  # value 값 불러옴 -> 2중 list 형태를 하나의 list 로 생성

    print((sorted(list(set(list2)))))  # landmark list 출력 ( 중복 제거 )

    j = ""
    for i in sorted(list(set(list2))):
        print(i)
        j = j + str(i) + " "  # 하나의 str 으로 만듬 => predict 할 때 사용
        landmark_list.append(i)  # landmark list 추가

    print(j)  # landmark num 나열

    return landmark_list


def load_xlsx(path: str):
    """
    on3d, on3ds 최신 엑셀을 읽어와 landmark_num : key, landmark_name : value 형태의 dict 로 저장
    """
    wb = load_workbook(path,
                       read_only=False,  # 읽기 전용(읽기 전용에 최적화 되어 파일을 불러온다)
                       data_only=False,  # False 면 셀안 공식을 가져오고 True 면 공식 적용된 값만을 불러온다.
                       )
    ws = wb.active  # 첫 번째 시트
    landmark_num = ''
    landmark_name = ''
    landmark_num_name_dict = {}

    row_max = ws.max_row  # 제일 마지막 행 저장

    get_cells = ws['A5': f'C{row_max}']  # 출력할 cell 범위 지정 => row_max 까지 사용 하지는 않지만 범위를 정해야 하기 때문에 사용
    for row in get_cells:

        for cell in row:
            if cell.value == 'Next Id':  # next id 가 있으면
                return landmark_num_name_dict

            elif 'B' in cell.coordinate:  # column = B ( landmark_num )     cell.coordinate => 셀 위치 출력
                landmark_num = cell.value

            elif 'C' in cell.coordinate:  # column = C ( landmark_name )
                landmark_name = cell.value

        landmark_num_name_dict[landmark_num] = landmark_name  # dict 저장

    print(landmark_num_name_dict)
    return landmark_num_name_dict


def compare_json_xlsx(json_land_list: list, xlsx_land_dict: dict):
    """
        set 을 통해서 차집합 각각 구해서 sorted
    """
    xlsx_land_list = list(xlsx_land_dict.keys())

    only_json = sorted(list(set(json_land_list) - set(xlsx_land_list)))
    only_xlsx = sorted(list(set(xlsx_land_list) - set(json_land_list)))

    print(f'\n기존 AI landmark 학습에서 제거할 landmark list: {only_json} <= 이 부분만 점검 사항에 추가하면 됨. \nAI 학습에 포함될 필요가 없는 landmark list : {only_xlsx}')


if __name__ == "__main__":
    json_landmark = load_json(json_path)
    xlsx_dict = load_xlsx(xlsx_path)
    compare_json_xlsx(json_landmark, xlsx_dict)
