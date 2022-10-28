import json
import openpyxl

# json 파일 오픈, key,value 나눔
from openpyxl.styles import Border, borders, PatternFill, Font, Alignment

json_loc = r"C:\Users\3DONS\Desktop\grouping_json\on3d-on3ds-on3d\temp_group_points_preShin.json"  # 변환?정리? 할 json 파일 경로
on3d_xlsx = r'C:/Users/3DONS/Desktop/수정Dr_Cho_Analysis_220906(foramen 추가).xlsx'  # on3d landmark 파일 경로
on3ds_xlsx = r'C:/Users/3DONS/Desktop/수정Dr_Cho_Analysis_ON3D_S_220906(foramen 추가).xlsx'  # on3ds landmark 파일 경로
arrange_xlsx = r'C:\woo_project\landmark_name/group_num_name.xlsx'  # 저장할 xlsx 경로, 이름
delete_group = ['TMJ_1', 'TMJ_2', 'TMJ_3', 'Cranial_Base']  # 제거 할 landmark group name

''' 랜드마크 xlsx 은 landmark 가 적혀 있는 표를 제외 하고 전부 삭제 해야 한다.'''

with open(json_loc, "r") as file:  # 변환할 json 파일
    data = json.load(file)

on3d_landmark_num = list(data.values())
on3d_landmark_name = []
on3d_group_name = list(data.keys())

on3d = {}
on3d_s = {}

# 기존 엑셀 open
wb = openpyxl.load_workbook(filename=on3d_xlsx)
ws = wb['Landmark']

# 기존 엑셀 데이터 번호와 json 번호 비교 후
# json value 값에 landmark 이름 넣음
for i in range(len(on3d_landmark_num)):
    temp_name = []
    for j in range(len(on3d_landmark_num[i])):
        for column in range(1, 186):  # 범위는 직접 확인 마지막 landmark num 행 +1
            if ws[f'B{column}'].value == on3d_landmark_num[i][j]:
                temp_name.append(ws[f'C{column}'].value)
                on3d[ws[f'B{column}'].value] = ws[f'C{column}'].value  # on3d value,name
    on3d_landmark_name.append(temp_name)
wb.close()

# s용 엑셀 open
wb = openpyxl.load_workbook(filename=on3ds_xlsx)
ws_2 = wb['Landmark']

# 위에서 바꾼 value 값과 비교 후 번호 대입
for key, value in on3d.items():
    for column in range(1, 146):
        if ws_2[f'C{column}'].value == value:
            on3d_s[ws_2[f'C{column}'].value] = ws_2[f'B{column}'].value  # name-value
        else:
            pass

wb = openpyxl.Workbook()
wb.save(arrange_xlsx)
wb = openpyxl.load_workbook(arrange_xlsx)
ws = wb['Sheet']
row = 4  # 시작 row

landmark_name = []  # 작성된 landmark list
x = []  # 판별 리스트
overlap_name = []  # 중복된 번호

for i in range(len(on3d_landmark_num)):  # xlsx 에 데이터 입력
    for j in range(len(on3d_landmark_num[i])):
        if on3d_group_name[i] in delete_group:
            pass
        else:
            ws.cell(row, 1).value = on3d_group_name[i]
            ws.cell(row, 5).value = on3d_group_name[i]
            ws.cell(row, 2).value = on3d_s.get(on3d.get(on3d_landmark_num[i][j], 'NONE'))
            ws.cell(row, 3).value = on3d.get(on3d_landmark_num[i][j], 'NONE')  # ~.get(a, b) dict 안에 앞에 (a)키가 없으면 (b)뒤에 출력
            ws.cell(row, 6).value = on3d_landmark_num[i][j]
            ws.cell(row, 7).value = on3d.get(on3d_landmark_num[i][j], 'NONE')
            landmark_name.append(on3d.get(on3d_landmark_num[i][j], 'NONE'))
            row += 1

# 중복 요소 검사
for k in landmark_name:
    if k not in x:
        x.append(k)
    else:
        if k not in overlap_name:
            overlap_name.append(k)    # 중복 값 append

for i in overlap_name:  # 중복 요소에 색상 입력
    for j in range(4, ws.max_row + 1):
        if i == ws.cell(j, 3).value:
            for k in range(3):
                ws.cell(j, k + 1).fill = PatternFill(start_color='ffffb3', end_color='ffffb3', fill_type='solid')
                ws.cell(j, k + 5).fill = PatternFill(start_color='ffffb3', end_color='ffffb3', fill_type='solid')

ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)  # 셀 병합
ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=7)
ws.cell(row=3, column=1).font = Font(bold=True)  # 텍스트 굵게
ws.cell(row=3, column=5).font = Font(bold=True)
ws.column_dimensions['A'].width = 20    # 셀 사이즈
ws.column_dimensions['E'].width = 20
ws.column_dimensions['C'].width = 25
ws.column_dimensions['G'].width = 25

for i in range(3, ws.max_row + 1):    # 시작 부터 마지막
    for j in range(3):
        ws.cell(i, j + 1).border = Border(left=borders.Side(style='thin'),  # 테두리
                                          right=borders.Side(style='thin'),
                                          top=borders.Side(style='thin'),
                                          bottom=borders.Side(style='thin'))
        ws.cell(i, j + 5).border = Border(left=borders.Side(style='thin'),
                                          right=borders.Side(style='thin'),
                                          top=borders.Side(style='thin'),
                                          bottom=borders.Side(style='thin'))
        ws.cell(i, j + 1).alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
        ws.cell(i, j + 5).alignment = Alignment(horizontal='center', vertical='center')

ws['A3'] = 'on3d_s group 번호, train 용'
ws['E3'] = 'on3d_s group 번호를 on3d로 변환한 group 번호, exporter 용'

wb.save(r'C:\woo_project\preprocessing/group_num_name.xlsx')
