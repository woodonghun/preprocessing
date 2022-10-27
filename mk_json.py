import json

import openpyxl

xlsx_loc = 'C:/Users/3DONS/Desktop/group_num_name_update04.xlsx'  # 정리 xlsx 경로
create_json_loc = 'C:/Users/3DONS/Desktop/1'  # 생성할 json 파일 폴더 경로


def make_json(landmark_dict: dict, loc: str):  # json 파일 생성
    with open(loc, 'w') as outfile:
        json.dump(landmark_dict, outfile)


def open_xlsx(loc: str, sheet_name: str):    # sheet_name : 시트이름 입력
    wb = openpyxl.load_workbook(loc)
    ws = wb[sheet_name]

    key = []
    on3ds_value = []
    on3ds_landmark_num = []
    on3d_landmark_num = []
    on3d_value = []
    on3d_json = {}
    on3ds_json = {}
    start_row = 4

    for i in range(start_row, ws.max_row + 1):  # 시작 row, 끝 row 로 key, value 값 추가
        key.append(ws.cell(i, 1).value)
        on3ds_value.append(ws.cell(i, 2).value)
        on3d_value.append(ws.cell(i, 6).value)

    for j in range(len(key)):    # dict 생성
        if j+1 == len(key):
            on3ds_landmark_num.append(on3ds_value[j])  # 마지막 key 값 일 때 따로 추가
            on3d_landmark_num.append(on3d_value[j])
            on3ds_json[key[j]] = on3ds_landmark_num
            on3d_json[key[j]] = on3d_landmark_num
            break
        elif key[j] == key[j + 1]:  # 뒤에 landmark name 과 같을 때 각각의 list 에 value 추가함.
            on3ds_landmark_num.append(on3ds_value[j])
            on3d_landmark_num.append(on3d_value[j])

        elif key[j] != key[j + 1]:
            on3ds_landmark_num.append(on3ds_value[j])  # 뒤에 landmark name 과 다를 때 list 에 value 추가 한 뒤 dict 로 생성 함.
            on3d_landmark_num.append(on3d_value[j])
            on3ds_json[key[j]] = on3ds_landmark_num
            on3d_json[key[j]] = on3d_landmark_num

            on3ds_landmark_num = []
            on3d_landmark_num = []

    return on3d_json, on3ds_json  # 각각의 dict 반환


def split_group_point(json_loc: str):  # landmark , 제거
    with open(json_loc, 'r') as t:
        json_data = json.load(t)

    value = list(json_data.values())
    key = list(json_data.keys())

    with open(f'{create_json_loc}/landmark.txt', 'a') as f:
        for i in range(len(key)):
            f.write(key[i] + ' ')    # key 입력
            for j in range(len(value[i])):
                f.write(str(value[i][j]) + ' ')    # value 입력
            f.write('\n')
        value = sum(value, [])
        value = list(set(value))    # list 중복 제거 , 나열
        for j in range(len(value)):
            f.write(str(value[j]) + ' ')


on3d, on3ds = open_xlsx(xlsx_loc, '정리')
make_json(on3d, f'{create_json_loc}/on3d.json')
make_json(on3ds, f'{create_json_loc}/on3ds.json')
split_group_point(f'{create_json_loc}/on3d.json')
