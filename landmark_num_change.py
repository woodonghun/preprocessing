import os

import openpyxl

xlsx_loc = 'C:/Users/3DONS/Desktop/group_num_name_update04.xlsx'  # on3d, on3d_s 번호 정리한 xlsx
landmark_loc = r'C:\Users\3DONS\Desktop\ON3DS_landmarks_Number_convert\on3d_landmarks'  # landmark 가 저장된 폴더
on3d_s_landmark_loc = r'C:\Users\3DONS\Desktop\ON3DS_landmarks_Number_convert\New_numbering_landmarks'  # 새로운 landmark 를 저장할 폴더

# landmarks_on3d_s_final_result 사용하면 됨
# landmarks_on3d_s, landmarks_on3d_s_final_result 폴더 2개가 생성됨.
# landmarks_on3d_s 는 단순히 번호만 변환한 것
# landmarks_on3d_s_final_result 에서는 번호 변환 및 , 중복 제거
# on3d -> on3ds 변환 과정중 없는 landmark 는 -99999.00 으로 변환
# 중복 번호는 landmark 좌표가 있는 번호를 남겨 두고 -99999만 제거 ( landmark 좌표가 같이 겹치는 게 생길 경우에는 exporter 에서 landmark 좌표를 잘못 입력)

os.mkdir(f'{on3d_s_landmark_loc}/landmarks_on3d_s')
os.mkdir(f'{on3d_s_landmark_loc}/landmarks_on3d_s_final_result')

file_list = os.listdir(landmark_loc)  # landmark 폴더의 txt 파일 리스트

wb = openpyxl.load_workbook(xlsx_loc)
ws = wb['정리']

on3d_s_num = []
on3d_s_landmark = []
on3d_num = []
on3d_landmark = []

for column in range(4, 124):  # 정리된 xlsx 에서 landmark 값 가지고 옴
    on3d_s_num.append(str(ws[f'B{column}'].value))
    on3d_s_landmark.append(ws[f'C{column}'].value)
    on3d_num.append(str(ws[f'F{column}'].value))
    on3d_landmark.append(ws[f'G{column}'].value)

on3d = dict(zip(on3d_landmark, on3d_num))  # key : landmark_name, value : landmark_number
on3d_s = dict(zip(on3d_s_landmark, on3d_s_num))  # dict 완성, 동일한 key 값이 있으면 자동 으로 제거됨

total_landmark = []
new_landmark = []
new_landmark_list = []
not_use = []
count = 1

for j in range(len(file_list)):
    # txt 읽고 변환
    with open(f"{landmark_loc}/{file_list[j]}", "r") as f:  # txt 읽기
        for line in f:  # 한줄씩 읽기
            landmark = (line.split(','))  # list 로 만듬 [landmark_number,x,y,z]
            for key, value in on3d.items():  # dict on3d 를 key, value for 구동
                if str(value) == landmark[0]:  # on3d 번호 == landmark 번호
                    landmark[0] = str(on3d_s[key])  # landmark 번호 -> on3d_s 번호

                    break  # break 안쓰면 다시 반복되서 꼭 써야함.

                # on3d.items 의 index 와 count 의 값은 동일하게 1씩 증가
                # index 와 count 가 동일하고, landmark 와 dict 의 value 값이 다르면 사용 하지 않은게 됨
                elif str(value) != landmark[0] and count >= len(on3d.items()):
                    landmark[1] = '-99999.00'
                    landmark[2] = '-99999.00'
                    landmark[3] = '-99999.00\n'
                    break
                count += 1
            result = ','.join(s for s in landmark)  # str 으로 변경
            with open(f'{on3d_s_landmark_loc}/landmarks_on3d_s/{file_list[j]}', 'a') as t:  # 파일 생성
                t.write(result)

            count = 1

    # 변환된 txt 읽고 중복 제거, 비어 있는 값 입력
    with open(f"{on3d_s_landmark_loc}/landmarks_on3d_s/{file_list[j]}", "r") as f:  # txt 읽기
        for line in f:  # 한줄씩 읽기
            landmark = (line.split(','))
            total_landmark.append(landmark)
    # 변환된 txt 내용 한줄씩 읽고 저장 후 sort
    # 중복 되는 번호가 순서대로 나열됨
    total_landmark.sort()

    for i in total_landmark:
        i[0] = str(i[0])

    for z in range(len(total_landmark)):
        if z == len(total_landmark) - 1:
            new_landmark.append(total_landmark[z])  # 마지막 index 로가면 index over 되서 바로 추가. 마지막 번호 뒤에는 아무것도 없기 때문에 중복 x
            break
        new_landmark.append(total_landmark[z])  # list 하나씩 추가
        if total_landmark[z][0] == total_landmark[z + 1][0]:
            if total_landmark[z][1] == '-99999.00':  # 경우의 수 3가지 , 앞 -99999, 뒤 -99999, 둘다 -99999
                new_landmark[z] = ''  # ''으로 변경
            elif total_landmark[z + 1][z] == '-99999.00':
                new_landmark[z] = ''
            elif total_landmark[z][1] == '-99999.00' and total_landmark[z + 1][z] == '-99999.00':
                new_landmark[z] = ''

    for g in range(len(new_landmark)):
        if new_landmark[g] == '':
            pass
        else:
            new_landmark_list.append(new_landmark[g][0])  # '' 값을 제외 하고 추가, 즉 사용한 landmark 의 list
    complement = list(set(on3d_s_num) - set(new_landmark_list))  # 사용 안한 on3ds landmark = (on3ds landmark) - (사용한 landmark)
    for b in range(len(complement)):
        not_use.append([str(complement[b]), '-99999.00', '-99999.00', '-99999.00\n'])  # 사용 안한 on3ds landmark 에 -99999 추가, 2중 list 제작
    insert_landmark = new_landmark + not_use  # 총 on3ds landmark

    for m in insert_landmark:
        m[0] = int(m[0])
    insert_landmark.sort()
    for m in insert_landmark:
        m[0] = str(m[0])

    for c in insert_landmark:
        if c == '':  # '' 이 포함 되어 있음, pass ( 마지막 txt 줄 끝에 '' 이 포함될수도있음)
            pass
        else:
            result = ','.join(s for s in c)  # str 으로 변경, 나머지 write
            with open(f'{on3d_s_landmark_loc}/landmarks_on3d_s_final_result/{file_list[j]}', 'a') as t:  # 파일 생성
                t.write(result)
    # print(insert_landmark)
    count = 1
    new_landmark = []
    total_landmark = []
    save_landmark = []
    new_landmark_list = []
    not_use = []
